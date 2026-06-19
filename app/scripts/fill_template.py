#!/usr/bin/env python3
"""
PI Agent - Excel Template Filler
혁신과제 상세정의서(PDF)에서 추출한 세부과제 데이터(JSON)를
PI 검토 결과서 Template(xlsx) 형식에 맞춰 채워 넣는다.

사용법:
  python fill_template.py --template <Template.xlsx> --data <data.json> --out <output.xlsx>

JSON 스키마는 pi_data.json 참조.
"""
import argparse, json, copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 데이터 행 컬럼 매핑 (Template 4행 헤더 기준)
COLS = {
    "domain": "A", "task_id": "B", "task_name": "C", "task_overview": "D",
    "sub_id": "E", "sub_name": "F", "definition": "G", "process_id": "H",
    "조선": "I", "해양": "J", "특수": "K", "미포": "L",
    "HD한조": "M", "HHIP": "N", "HVS": "O", "삼호": "P",
    "owner": "Q", "solution": "R", "effect_q": "S", "effect_n": "T",
}
DATA_START_ROW = 5  # 헤더는 1~4행
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_COLS = {"C", "D", "F", "G", "R", "S", "T"}   # 긴 텍스트 → 줄바꿈/상단정렬
CENTER_COLS = {"A", "B", "E", "H", "I", "J", "K", "L", "M", "N", "O", "P"}
COL_WIDTH = {"A": 12, "B": 6, "C": 24, "D": 34, "E": 8, "F": 26, "G": 40,
             "H": 12, "I": 5, "J": 5, "K": 5, "L": 5, "M": 7, "N": 6, "O": 6,
             "P": 6, "Q": 22, "R": 18, "S": 38, "T": 14}


def style_cell(cell, col):
    cell.font = Font(name="맑은 고딕", size=10)
    cell.border = BORDER
    if col in WRAP_COLS:
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    elif col in CENTER_COLS:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    data = json.load(open(args.data, encoding="utf-8"))
    wb = load_workbook(args.template)
    ws = wb["Sheet1"]

    # 기존 예시(5행 이하) 정리
    if ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)

    bu = data["owner_bu_map"]
    rows = []                       # (row_idx, task_id) 기록 → 그룹 병합용
    r = DATA_START_ROW
    for task in data["tasks"]:
        for st in task["subtasks"]:
            rec = {
                "domain": data["domain"], "task_id": task["task_id"],
                "task_name": task["task_name"], "task_overview": task["task_overview"],
                "sub_id": st["sub_id"], "sub_name": st["sub_name"],
                "definition": st["definition"], "process_id": "",
                "owner": data["owner_text"], "solution": st.get("solution", ""),
                "effect_q": task["effect_q"], "effect_n": task.get("effect_n", ""),
                **bu,
            }
            for key, col in COLS.items():
                c = ws[f"{col}{r}"]
                c.value = rec.get(key, "")
                style_cell(c, col)
            ws.row_dimensions[r].height = 78
            rows.append((r, task["task_id"]))
            r += 1

    # 세로 병합: 도메인(A)은 전체, 과제ID/명/개요·오너는 과제 단위
    last = rows[-1][0]
    ws.merge_cells(f"A{DATA_START_ROW}:A{last}")
    groups = {}
    for ridx, tid in rows:
        groups.setdefault(tid, []).append(ridx)
    for tid, idxs in groups.items():
        if len(idxs) > 1:
            for col in ("B", "C", "D", "Q"):
                ws.merge_cells(f"{col}{idxs[0]}:{col}{idxs[-1]}")

    for col, w in COL_WIDTH.items():
        ws.column_dimensions[col].width = w

    wb.save(args.out)
    print(f"saved: {args.out}  ({len(rows)} rows)")


if __name__ == "__main__":
    main()
